import re
import io
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

# ── Colours (from app.py) ───────────────────────────────────────────────────
ACCENT = "#00e5a0"
ACCENT2 = "#0ea5e9"
PURPLE = "#a78bfa"
WARN = "#f59e0b"
DANGER = "#ef4444"
MUTED = "#64748b"
TEXT = "#e2e8f0"
DARK = "#0d1117"
BG = DARK
CARD = "#111720"
BORDER = "#1e2a38"

# Configure matplotlib with the dark layout color palette
plt.rcParams.update(
    {
        "figure.facecolor": CARD,
        "axes.facecolor": DARK,
        "axes.edgecolor": BORDER,
        "axes.labelcolor": MUTED,
        "axes.titlecolor": TEXT,
        "xtick.color": MUTED,
        "ytick.color": MUTED,
        "text.color": TEXT,
        "grid.color": BORDER,
        "grid.linewidth": 0.5,
        "font.family": "monospace",
        "font.size": 9,
    }
)

# 1. COLUMN_MAP & _COLUMN_PATTERNS
COLUMN_MAP = {
    # ── Exact columns from the real data ──────────────────────────────────────
    r"#": "row_number",
    r"paper.?code": "voucher_id",
    r"dispensing.?date": "visit_date",
    r"patient.?name": "patient_name",
    r"patient.?type": "patient_type",
    r"gender": "gender",
    r"is.?newborn": "is_newborn",
    r"rama.?number": "patient_id",
    r"practitioner.?name": "doctor_name",
    r"practitioner.?type": "doctor_type",
    r"total.?cost": "amount",
    r"patient.?co.?payment": "patient_copay",
    r"insurance.?co.?payment": "insurance_copay",
    r"medicine.?cost": "medicine_cost",
    # ── Generic fallbacks ─────────────────────────────────────────────────────
    r"patient.?(id|no|num|number|code)?": "patient_id",
    r"pat.?id|pid": "patient_id",
    r"doctor.?(id|no|num|code)?": "doctor_id",
    r"(doctor|dr|physician|prescriber).?name": "doctor_name",
    r"doc.?id|did": "doctor_id",
    r"prescriber": "doctor_name",
    r"(visit|service|rx|voucher).?date": "visit_date",
    r"date.?(of.?)?(visit|service|dispensing)?": "visit_date",
    r"date": "visit_date",
    r"(pharmacy|facility|clinic|hospital|branch).?(name|id|code)?": "facility",
    r"(drug|medicine|medication|item|product).?(name|description|desc)?": "drug_name",
    r"(drug|medicine|medication|item|product).?(code|id)?": "drug_code",
    r"(amount|cost|price|value|total|charge)": "amount",
    r"quantity|qty": "quantity",
    r"(diagnosis|diag|icd|condition)": "diagnosis",
    r"(voucher|claim|ref|reference).?(no|number|id|code)?": "voucher_id",
}

_COLUMN_PATTERNS = [
    (re.compile(pattern), target) for pattern, target in COLUMN_MAP.items()
]

# 2. load_and_process()
def load_and_process(file_bytes: bytes, filename: str, rapid_days: int):
    fname = filename.lower()
    if fname.endswith(".csv"):
        df = pd.read_csv(
            io.BytesIO(file_bytes),
            encoding="utf-8",
            on_bad_lines="skip",
            dtype_backend="pyarrow",
        )
    elif fname.endswith((".xlsx", ".xls")):
        df = pd.read_excel(io.BytesIO(file_bytes), dtype_backend="pyarrow")
    elif fname.endswith(".ods"):
        df = pd.read_excel(
            io.BytesIO(file_bytes), engine="odf", dtype_backend="pyarrow"
        )
    else:
        raise ValueError("Unsupported file type. Use CSV, XLSX, XLS, or ODS.")

    # Normalise column names
    renamed, used = {}, {}
    for col in df.columns:
        key = re.sub(r"[^a-z0-9]", "_", col.lower().strip())
        key = re.sub(r"_+", "_", key).strip("_")
        matched = False
        for pattern, target in _COLUMN_PATTERNS:
            if pattern.fullmatch(key):
                if target not in used:
                    renamed[col] = target
                    used[target] = col
                    matched = True
                break
        if not matched:
            renamed[col] = key
    df = df.rename(columns=renamed)

    # Parse dates
    if "visit_date" in df.columns:
        df["visit_date"] = pd.to_datetime(df["visit_date"], errors="coerce")
    else:
        for col in df.columns:
            _dstr = str(df[col].dtype)
            if _dstr == "object" or "string" in _dstr or "large_string" in _dstr:
                try:
                    parsed = pd.to_datetime(df[col], errors="coerce")
                    if parsed.notna().sum() > len(df) * 0.5:
                        df["visit_date"] = parsed
                        break
                except Exception:
                    pass

    # Summary stats
    s = {"total_rows": len(df), "columns": list(df.columns)}
    id_col = (
        "patient_id"
        if "patient_id" in df.columns
        else "patient_name" if "patient_name" in df.columns else None
    )
    if id_col:
        vc = df[id_col].value_counts()
        s["patient_col"] = id_col
        s["unique_patients"] = int(df[id_col].nunique())
        s["repeat_patients"] = int((vc > 1).sum())
        s["max_visits"] = int(vc.max())
        s["top_patients"] = vc.head(15).rename_axis("id").reset_index(name="visits")

    dcol = (
        "doctor_name"
        if "doctor_name" in df.columns
        else "doctor_id" if "doctor_id" in df.columns else None
    )
    if dcol:
        dvc = df[dcol].value_counts()
        s["unique_doctors"] = int(df[dcol].nunique())
        s["top_doctors"] = dvc.head(15).rename_axis("doctor").reset_index(name="visits")
        s["doctor_col"] = dcol

    if "visit_date" in df.columns:
        v = df["visit_date"].dropna()
        if len(v):
            s["date_min"] = str(v.min().date())
            s["date_max"] = str(v.max().date())

    if "facility" in df.columns:
        fvc = df["facility"].value_counts()
        s["unique_facilities"] = int(df["facility"].nunique())
        s["top_facilities"] = (
            fvc.head(10).rename_axis("name").reset_index(name="visits")
        )

    for amt_col in ["amount", "medicine_cost", "insurance_copay", "patient_copay"]:
        if amt_col in df.columns:
            df[amt_col] = pd.to_numeric(df[amt_col], errors="coerce")

    if "amount" in df.columns:
        s["total_amount"] = round(float(df["amount"].sum()), 2)
        s["avg_amount"] = round(float(df["amount"].mean()), 2)

    # Repeat visits
    repeat_groups, repeat_detail = [], pd.DataFrame()
    if id_col:
        vc2 = df[id_col].value_counts()
        repeat_ids = vc2[vc2 > 1].index.tolist()
        rdf = df[df[id_col].isin(repeat_ids)].copy()
        if "visit_date" in rdf.columns:
            rdf = rdf.sort_values([id_col, "visit_date"])
        repeat_detail = rdf.head(500)

        has_name = "patient_name" in rdf.columns and id_col != "patient_name"
        has_date = "visit_date" in rdf.columns
        top_rdf = rdf[rdf[id_col].isin(repeat_ids[:300])]
        for pid, grp in top_rdf.groupby(id_col, sort=False):
            entry = {id_col: str(pid), "visits": int(len(grp))}
            if has_name:
                entry["patient_name"] = str(grp["patient_name"].iloc[0])
            if has_date:
                dates = grp["visit_date"].dropna().sort_values()
                entry["dates"] = ", ".join(str(d.date()) for d in dates if pd.notna(d))
            repeat_groups.append(entry)
        repeat_groups.sort(key=lambda x: x["visits"], reverse=True)

    # 3. Rapid Revisit Engine
    rapid = []
    if id_col and "visit_date" in df.columns:
        cols = [id_col, "visit_date"]
        if "patient_name" in df.columns and id_col != "patient_name":
            cols.append("patient_name")
        if dcol:
            cols.append(dcol)
        sub = (
            df[cols]
            .dropna(subset=[id_col, "visit_date"])
            .sort_values([id_col, "visit_date"])
        )

        sub["_prev_date"] = sub.groupby(id_col)["visit_date"].shift(1)
        sub["_days_diff"] = (sub["visit_date"] - sub["_prev_date"]).dt.days

        rapid_mask = (sub["_days_diff"] > 0) & (sub["_days_diff"] <= rapid_days)
        rapid_df = sub[rapid_mask].copy()

        if len(rapid_df) > 0:
            rapid_dict = {
                "patient_id": rapid_df[id_col].astype(str).tolist(),
                "patient_name": (
                    rapid_df["patient_name"].astype(str).tolist()
                    if "patient_name" in rapid_df.columns
                    else rapid_df[id_col].astype(str).tolist()
                ),
                "visit_1": rapid_df["_prev_date"].dt.strftime("%Y-%m-%d").tolist(),
                "visit_2": rapid_df["visit_date"].dt.strftime("%Y-%m-%d").tolist(),
                "days_apart": rapid_df["_days_diff"].astype(int).tolist(),
            }
            if dcol and dcol in rapid_df.columns:
                rapid_dict["doctor"] = rapid_df[dcol].astype(str).tolist()
            else:
                rapid_dict["doctor"] = ["—"] * len(rapid_df)

            rapid = [dict(zip(rapid_dict.keys(), row)) for row in zip(*rapid_dict.values())]
            rapid.sort(key=lambda x: x["days_apart"])

    return df, renamed, s, repeat_groups, repeat_detail, rapid

# 4. Chart Generation Buffers
def get_buffer(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", transparent=True)
    buf.seek(0)
    plt.close(fig)
    return buf

def hbar_chart_buf(labels, values, color, title, xlabel):
    if not values:
        return None
    max_val = max(values) or 1
    fig, ax = plt.subplots(figsize=(7, max(2.5, len(labels) * 0.42)))
    bars = ax.barh(
        labels[::-1],
        values[::-1],
        color=color if isinstance(color, list) else [color] * len(labels),
        height=0.65,
    )
    for bar, val in zip(bars, values[::-1]):
        ax.text(
            bar.get_width() + max_val * 0.01,
            bar.get_y() + bar.get_height() / 2,
            str(val),
            va="center",
            color=TEXT,
            fontsize=8,
        )
    ax.set_xlabel(xlabel)
    ax.set_title(title, fontsize=11, fontweight="bold", color=TEXT, pad=10)
    ax.set_xlim(0, max_val * 1.2)
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(axis="x", alpha=0.3)
    fig.tight_layout()
    return get_buffer(fig)

def time_series_chart_buf(df):
    if "visit_date" not in df.columns:
        return None
    s = df["visit_date"].dropna()
    if len(s) < 2:
        return None
    monthly = s.dt.to_period("M").value_counts().sort_index()
    dates, vals = [str(p) for p in monthly.index], monthly.values
    fig, ax = plt.subplots(figsize=(10, 3))
    ax.fill_between(range(len(vals)), vals, alpha=0.2, color=ACCENT)
    ax.plot(range(len(vals)), vals, color=ACCENT, linewidth=2, marker="o", markersize=4)
    step = max(1, len(dates) // 12)
    ax.set_xticks(range(0, len(dates), step))
    ax.set_xticklabels(dates[::step], rotation=45, ha="right", fontsize=8)
    ax.set_ylabel("Visits")
    ax.set_title(
        "Monthly Visit Volume", fontsize=11, fontweight="bold", color=TEXT, pad=10
    )
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    return get_buffer(fig)

def rapid_histogram_buf(rapid):
    if not rapid:
        return None
    days = [r["days_apart"] for r in rapid]
    fig, ax = plt.subplots(figsize=(6, 3))
    bins = list(range(1, max(days) + 2))
    _, bins_out, patches = ax.hist(
        days, bins=bins, color=WARN, edgecolor=CARD, rwidth=0.8
    )
    for patch, left in zip(patches, bins_out):
        if left <= 2:
            patch.set_facecolor(DANGER)
    ax.set_xlabel("Days Between Visits")
    ax.set_ylabel("Cases")
    ax.set_title(
        "Rapid Revisit Distribution", fontsize=11, fontweight="bold", color=TEXT, pad=10
    )
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(axis="y", alpha=0.3)
    ax.legend(
        handles=[
            mpatches.Patch(color=DANGER, label="≤2 days"),
            mpatches.Patch(color=WARN, label="3+ days"),
        ],
        fontsize=8,
        facecolor=CARD,
        edgecolor=BORDER,
        labelcolor=TEXT,
    )
    fig.tight_layout()
    return get_buffer(fig)

import networkx as nx
import difflib
import math
from collections import defaultdict as _dd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Network Intelligence ──────────────────────────────────────────────────

def build_network_data(df, col_a='doctor_name', col_b='patient_id', max_nodes=100, min_weight=1):
    if col_a not in df.columns or col_b not in df.columns:
        return {"nodes": [], "edges": []}

    sub = df[[col_a, col_b]].dropna()
    edge_counts = sub.groupby([col_a, col_b]).size().reset_index(name="weight")
    edge_counts = edge_counts[edge_counts["weight"] >= min_weight]

    G = nx.Graph()
    for _, row in edge_counts.iterrows():
        a, b, w = str(row[col_a]), str(row[col_b]), int(row["weight"])
        G.add_node(a, side="A")
        G.add_node(b, side="B")
        G.add_edge(a, b, weight=w)

    if len(G.nodes) > max_nodes:
        degree_list = sorted(list(G.degree()), key=lambda x: x[1], reverse=True)
        top_nodes = [n for n, _ in degree_list[:max_nodes]]
        G = G.subgraph(top_nodes).copy()

    degrees = dict(G.degree())
    nodes = []
    for n, data in G.nodes(data=True):
        is_a = data.get("side") == "A"
        deg = degrees.get(n, 1)
        nodes.append({
            "id": n,
            "label": str(n)[:20],
            "title": f"Connections: {deg}",
            "color": "#00e5a0" if is_a else "#0ea5e9",
            "shape": "diamond" if is_a else "dot",
            "size": max(10, min(30, 10 + deg * 2))
        })

    edges = []
    for u, v, data in G.edges(data=True):
        edges.append({
            "from": u,
            "to": v,
            "value": data.get("weight", 1),
            "color": {"color": "#1e2a38", "highlight": "#00e5a0"}
        })

    return {"nodes": nodes, "edges": edges}

# ── Fraud Detection & Fuzzy Matching ────────────────────────────────────────

def _toks(name):
    s = re.sub(r"[-_]", " ", str(name).lower())
    return set(re.sub(r"[^a-z0-9 ]", "", s).split())

def _match_score(a, b):
    ta, tb = _toks(a), _toks(b)
    if not ta or not tb: return 0.0
    shorter, longer = (ta, tb) if len(ta) <= len(tb) else (tb, ta)
    if shorter <= longer: return 0.9, "subset"
    ratio = difflib.SequenceMatcher(None, " ".join(sorted(ta)), " ".join(sorted(tb))).ratio()
    return ratio, "fuzzy"

def normalize_rama(val):
    s = re.sub(r"[^A-Z0-9]", "", str(val).upper())
    return s

def run_match(ph_work, fac_df, name_thresh=0.4, date_window=7, require_name=True):
    ph = ph_work.copy()
    fac = fac_df.copy()
    ph["_rn"] = ph["_rama"].apply(normalize_rama)
    fac["_rn"] = fac["_rama"].apply(normalize_rama)

    fac_index = _dd(list)
    for fr in fac.to_dict("records"):
        fac_index[fr["_rn"]].append(fr)

    results = []
    for pr in ph.to_dict("records"):
        candidates = fac_index.get(pr["_rn"], [])
        if not candidates:
            results.append({**pr, "status": "NO_RECORD", "confidence": 0.0})
            continue

        best_fr, best_conf, best_ns, best_days = None, -1.0, 0.0, None
        for fr in candidates:
            ns, _ = _match_score(pr["_name"], fr["_name"])
            if require_name and ns < name_thresh: continue

            days = None
            if pd.notna(pr["_date"]) and pd.notna(fr["_date"]):
                days = (pd.Timestamp(pr["_date"]) - pd.Timestamp(fr["_date"])).days

            if days is not None and -1 <= days <= date_window:
                day_prox = 1.0 - max(days, 0) / (date_window + 1)
                conf = round(0.4 * ns + 0.6 * day_prox, 3)
                if conf > best_conf:
                    best_conf, best_fr, best_ns, best_days = conf, fr, ns, days

        if best_fr:
            results.append({**pr, "status": "MATCHED", "confidence": best_conf, "fac_source": best_fr["_source"], "days_apart": best_days})
        else:
            results.append({**pr, "status": "UNLINKED", "confidence": 0.0})

    return pd.DataFrame(results)

# ── Excel Audit Export ─────────────────────────────────────────────────────

def generate_excel_audit_report(results_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fraud Audit"

    header_fill = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
    font_white = Font(name="Calibri", size=11, color="FFFFFF", bold=True)

    cols = list(results_df.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(1, ci, col)
        cell.fill = header_fill
        cell.font = font_white
        ws.column_dimensions[get_column_letter(ci)].width = 20

    ws.row_dimensions[1].height = 25

    for ri, row in enumerate(results_df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, str(val))
            cell.font = Font(name="Calibri", size=10)
        ws.row_dimensions[ri].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
